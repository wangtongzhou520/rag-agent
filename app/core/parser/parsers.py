"""M1/M2 本地文档解析器。"""

import csv
import json
import re
from io import BytesIO, StringIO
from typing import Protocol

import chardet
from bs4 import BeautifulSoup
from markdown_it import MarkdownIt
from openpyxl import load_workbook

from app.core.parser.models import (
    CodeBlock,
    HeadingBlock,
    ParagraphBlock,
    ParsedDocument,
    Provenance,
    TableBlock,
)
from app.framework.exceptions import ClientException, ServiceException


class DocumentParser(Protocol):
    name: str

    def parse_structured(
        self, data: bytes, mime: str, options: dict | None = None
    ) -> ParsedDocument: ...


def decode_text(data: bytes) -> str:
    encoding = (chardet.detect(data).get("encoding") or "utf-8").lower()
    try:
        return data.decode(encoding, errors="strict").lstrip("\ufeff")
    except (LookupError, UnicodeDecodeError):
        return data.decode("utf-8", errors="replace").lstrip("\ufeff")


class TextDocumentParser:
    name = "text"

    def parse_structured(
        self, data: bytes, mime: str, options: dict | None = None
    ) -> ParsedDocument:
        filename = (options or {}).get("sourceFile")
        text = decode_text(data)
        if mime == "text/html":
            text = BeautifulSoup(text, "html.parser").get_text("\n")
        elif mime == "application/json":
            try:
                text = json.dumps(json.loads(text), ensure_ascii=False, indent=2)
            except json.JSONDecodeError as exc:
                raise ClientException("JSON 文件格式错误") from exc
        paragraphs = [
            ParagraphBlock(part.strip(), Provenance(filename))
            for part in re.split(r"\n\s*\n", text)
            if part.strip()
        ]
        return ParsedDocument(tuple(paragraphs), {"parser": self.name, "mimeType": mime})


class MarkdownDocumentParser:
    name = "markdown"

    def __init__(self) -> None:
        self._markdown = MarkdownIt("commonmark", {"html": True})

    def parse_structured(
        self, data: bytes, mime: str, options: dict | None = None
    ) -> ParsedDocument:
        filename = (options or {}).get("sourceFile")
        provenance = Provenance(filename)
        tokens = self._markdown.parse(decode_text(data))
        blocks = []
        index = 0
        while index < len(tokens):
            token = tokens[index]
            if token.type == "heading_open" and index + 1 < len(tokens):
                blocks.append(
                    HeadingBlock(int(token.tag[1]), tokens[index + 1].content, provenance)
                )
                index += 3
                continue
            if token.type == "paragraph_open" and index + 1 < len(tokens):
                blocks.append(ParagraphBlock(tokens[index + 1].content, provenance))
                index += 3
                continue
            if token.type in {"fence", "code_block"}:
                blocks.append(
                    CodeBlock(token.content, token.info.strip() or None, provenance)
                )
            index += 1
        return ParsedDocument(
            tuple(blocks), {"parser": self.name, "mimeType": mime, "blocks": len(blocks)}
        )


class CsvDocumentParser:
    name = "csv"

    def parse_structured(
        self, data: bytes, mime: str, options: dict | None = None
    ) -> ParsedDocument:
        rows = list(csv.reader(StringIO(decode_text(data))))
        if not rows:
            return ParsedDocument((), {"parser": self.name, "mimeType": mime})
        width = max(len(row) for row in rows)
        normalized = [tuple(row + [""] * (width - len(row))) for row in rows]
        block = TableBlock(
            normalized[0],
            tuple(normalized[1:]),
            Provenance((options or {}).get("sourceFile")),
        )
        return ParsedDocument((block,), {"parser": self.name, "mimeType": mime})


class ExcelDocumentParser:
    name = "excel"

    def parse_structured(
        self, data: bytes, mime: str, options: dict | None = None
    ) -> ParsedDocument:
        try:
            workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            raise ClientException("Excel 文件格式错误") from exc
        filename = (options or {}).get("sourceFile")
        blocks = []
        for sheet in workbook.worksheets:
            if sheet.sheet_state != "visible":
                continue
            values = [
                tuple("" if value is None else str(value) for value in row)
                for row in sheet.iter_rows(values_only=True)
            ]
            values = [row for row in values if any(cell.strip() for cell in row)]
            if not values:
                continue
            provenance = Provenance(filename, sheet.title)
            blocks.append(HeadingBlock(1, sheet.title, provenance))
            blocks.append(TableBlock(values[0], tuple(values[1:]), provenance))
        workbook.close()
        return ParsedDocument(tuple(blocks), {"parser": self.name, "mimeType": mime})


class PdfDocumentParser:
    name = "pdf"

    def parse_structured(
        self, data: bytes, mime: str, options: dict | None = None
    ) -> ParsedDocument:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise ServiceException("PDF 解析组件 pypdf 未安装") from exc
        try:
            reader = PdfReader(BytesIO(data))
            blocks = tuple(
                ParagraphBlock(text, Provenance((options or {}).get("sourceFile")))
                for page in reader.pages
                if (text := (page.extract_text() or "").strip())
            )
        except Exception as exc:
            raise ClientException("PDF 文件格式错误") from exc
        return ParsedDocument(blocks, {"parser": self.name, "mimeType": mime})
